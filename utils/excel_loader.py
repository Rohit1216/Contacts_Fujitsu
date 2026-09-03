"""Reads the standard team workbook format:

    Name | Phone | Email | Designation | Location | PhotoUrl | LinkedinUrl

One sheet per team/slide-group. Header row required; column order doesn't
matter as long as the header names match (case-insensitive).
"""

import openpyxl

COLUMN_MAP = {
    "name": "name",
    "phone": "phone",
    "email": "email",
    "designation": "designation",
    "title": "designation",
    "location": "location",
    "photourl": "photo_url",
    "photo url": "photo_url",
    "linkedinurl": "linkedin_url",
    "linkedin url": "linkedin_url",
}


def load_workbook_people(path: str) -> dict:
    """Returns {sheet_name: [person_dict, ...]}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        fields = [COLUMN_MAP.get(h) for h in header]

        people = []
        for row in rows[1:]:
            record = {}
            for field_name, value in zip(fields, row):
                if field_name is None:
                    continue
                if isinstance(value, str):
                    value = value.strip() or None
                record[field_name] = value
            if record.get("name"):
                people.append(record)
        result[sheet_name] = people
    return result


def dedupe_by_linkedin(*people_lists):
    """Merge several sheets into one contact list, de-duplicating people who
    appear on more than one sheet (matched by LinkedIn URL, falling back to
    name) while preserving first-seen order."""
    seen = set()
    merged = []
    for people in people_lists:
        for p in people:
            key = (p.get("linkedin_url") or p.get("name", "")).lower()
            if key in seen:
                continue
            seen.add(key)
            merged.append(p)
    return merged
